import os
import pandas as pd
import openpyxl
import zipfile
from datetime import datetime
from typing import List, Tuple
import pytz
from app.schemas.report_schemas import FlatnessResponse, ReportItem, ReportStatistics, FlatnessData
from app.config.app_config import REPORTS_DIR
from app.utils.public_fun import safe_float_convert


def get_reports_list() -> List[dict]:
    """
    获取报告目录下的所有Excel文件列表
    """
    excel_dir_path = REPORTS_DIR

    # 确保目录存在
    if not os.path.exists(excel_dir_path):
        print(f"报告目录不存在: {excel_dir_path}")
        return []

    data = []
    for filename in os.listdir(excel_dir_path):
        file_path = os.path.join(excel_dir_path, filename)

        # 只处理Excel文件
        if os.path.isfile(file_path) and filename.endswith(('.xlsx', '.xls')):
            try:
                # 获取文件创建时间
                file_create_time = os.path.getctime(file_path)
                file_create_dt_naive = datetime.fromtimestamp(file_create_time)

                # 使用pytz设置时区
                tz = pytz.timezone('Asia/Shanghai')
                file_create_dt = tz.localize(file_create_dt_naive)

                # 获取文件大小
                file_size = os.path.getsize(file_path)

                # 添加到数据列表
                data.append({
                    'id': 0,  # 由于不依赖数据库，ID设为0
                    'file_name': filename,
                    'created_at': file_create_dt.strftime('%Y-%m-%d %H:%M:%S'),
                    'file_size': file_size,
                    'flatness_count': 0,  # 需要读取文件才能知道具体数据条数
                    'sort_timestamp': file_create_time  # 用于排序的时间戳
                })
            except Exception as e:
                print(f"处理文件 {filename} 时发生错误: {str(e)}")
                continue

    # 按创建时间倒序排列（后创建的在前面）
    data.sort(key=lambda x: x['sort_timestamp'], reverse=True)

    # 移除排序用的时间戳字段，只保留需要的字段
    for item in data:
        del item['sort_timestamp']

    return data


def read_flatness_measure_data(file_path: str, worksheet_name: str = 'Flatness') -> Tuple[pd.DataFrame, dict]:
    """
    读取平面度报告Excel文件中的测量数据
    """
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(filename=file_path, data_only=True)

        # 检查指定的工作表是否存在
        if worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            # 如果指定的工作表不存在，尝试使用默认的第一个工作表
            ws = wb.worksheets[0]
            print(f"警告: 工作表 '{worksheet_name}' 不存在，使用默认工作表 '{ws.title}'")

        # 读取报告基本信息
        raw_measure_time = ws["C2"].value
        raw_max_value = ws["B8"].value
        raw_min_value = ws["B9"].value
        raw_pv_value = ws["B10"].value
        raw_rms = ws["B11"].value
        
        report_info = {
            'measure_time': raw_measure_time,
            'max_value': safe_float_convert(raw_max_value),
            'min_value': safe_float_convert(raw_min_value),
            'pv_value': safe_float_convert(raw_pv_value),  # 这是峰峰值（P-V值）
            'rms': safe_float_convert(raw_rms)            # 这是RMS值
        }

        # 读取测量数据（从第29行开始，对应row_index=29）
        measure_data = []
        current_row = 29  # 数据起始行
        # 循环读取数据直到遇到空行
        while True:
            # 检查当前行是否有数据
            index_cell = ws[f"A{current_row}"]
            angle_cell = ws[f"B{current_row}"]
            a_cell = ws[f"C{current_row}"]

            # 如果角度单元格为空，说明已到数据末尾
            if angle_cell.value is None:
                break

            # 添加数据到列表
            measure_data.append({
                'Index': index_cell.value,  # 排序信息（行号-26）
                'Angle': angle_cell.value,
                'A': a_cell.value
            })

            current_row += 1

        # 转换为DataFrame
        df = pd.DataFrame(measure_data)

        # 关闭工作簿
        wb.close()

        return df, report_info

    except Exception as e:
        raise Exception(f"读取Excel文件时发生错误: {str(e)}")


def get_flatness_data_by_filename(filename: str, worksheet_name: str = 'Flatness') -> FlatnessResponse:
    """
    根据文件名获取平整度数据
    """
    file_path = os.path.join(REPORTS_DIR, filename)

    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    # 检查文件是否为Excel文件
    if not filename.endswith(('.xlsx', '.xls')):
        raise ValueError(f"文件不是有效的Excel文件: {filename}")

    # 读取Excel文件数据
    measure_df, report_info = read_flatness_measure_data(file_path, worksheet_name)

    if measure_df is None or report_info is None:
        raise Exception("无法读取Excel文件数据")

    # 使用从Excel文件中读取的measure_time作为报告创建时间
    report_time_str = report_info.get('measure_time')

    # 使用pytz设置时区
    tz = pytz.timezone('Asia/Shanghai')

    if report_time_str:
        try:
            # 尝试解析报告时间字符串
            if isinstance(report_time_str, datetime):
                report_created_at = report_time_str
            else:
                # 假设格式为"YYYY-MM-DD HH:MM"
                report_created_at = datetime.strptime(str(report_time_str), "%Y-%m-%d %H:%M")
                report_created_at = tz.localize(report_created_at)
        except (ValueError, TypeError):
            # 如果解析失败，使用文件创建时间作为备选
            file_create_time = os.path.getctime(file_path)
            file_create_dt_naive = datetime.fromtimestamp(file_create_time)
            report_created_at = tz.localize(file_create_dt_naive)
    else:
        # 如果没有measure_time，使用文件创建时间
        file_create_time = os.path.getctime(file_path)
        file_create_dt_naive = datetime.fromtimestamp(file_create_time)
        report_created_at = tz.localize(file_create_dt_naive)

    # 提取平面度值用于统计计算
    flatness_values = [float(row['A']) for _, row in measure_df.iterrows() if 'A' in row and row['A'] is not None]

    # 直接使用从Excel中读取的统计值，它们已经经过类型检查并设置默认值为0
    max_value = report_info.get('max_value', 0)
    min_value = report_info.get('min_value', 0)
    peak_to_peak = report_info.get('pv_value', 0)
    rms_value = report_info.get('rms', 0)

    # 构建返回数据
    report_item = ReportItem(
        id=0,  # 由于直接从文件读取，没有数据库ID
        file_name=filename,
        bladeId=report_info.get('Balde_ID', ''),
        created_at=report_created_at.strftime('%Y-%m-%d %H:%M:%S')
    )

    statistics = ReportStatistics(
        max_value=round(max_value, 6),
        min_value=round(min_value, 6),
        peak_to_peak=round(peak_to_peak, 6),
        rms_value=round(rms_value, 6),
        data_count=len(flatness_values)
    )

    # 这个是后面的孔和孔的平面度值
    flatness_data = [
        FlatnessData(
            holeIndex=int(row['Index']) if row['Index'] is not None else 0,  # 叶片孔的排序信息
            holeAngle=float(row['Angle']),
            flatness=float(row['A'])
        ) for _, row in measure_df.iterrows()
        if 'Index' in row and 'Angle' in row and 'A' in row
        and row['Index'] is not None and row['Angle'] is not None and row['A'] is not None
    ]

    return FlatnessResponse(
        report=report_item,
        statistics=statistics,
        flatness_data=flatness_data
    )


def get_available_worksheets(filename: str) -> dict:
    """
    获取Excel文件中可用的工作表名称
    """
    file_path = os.path.join(REPORTS_DIR, filename)

    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    # 检查文件是否为Excel文件
    if not filename.endswith(('.xlsx', '.xls')):
        raise ValueError(f"文件不是有效的Excel文件: {filename}")

    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(filename=file_path, data_only=True)
        
        # 获取所有工作表名称
        worksheets = wb.sheetnames
        
        # 关闭工作簿
        wb.close()
        
        return {
            'worksheets': worksheets,
            'has_flatness': 'Flatness' in worksheets,
            'has_flatness_before': 'FlatnessBefore' in worksheets,
            'has_blade_result': 'BladeResult' in worksheets
        }
    except zipfile.BadZipFile:
        raise Exception(f"文件不是有效的Excel文件: 文件可能已损坏或不是.xlsx格式")
    except Exception as e:
        raise Exception(f"读取Excel文件工作表时发生错误: {str(e)}")


def get_blade_result_data(filename: str) -> dict:
    """
    从Excel文件的BladeResult工作表中获取加工信息
    """
    file_path = os.path.join(REPORTS_DIR, filename)
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(filename=file_path, data_only=True)
        
        # 检查是否存在BladeResult工作表
        if 'BladeResult' not in wb.sheetnames:
            wb.close()
            return {}  # 如果没有BladeResult工作表，则返回空字典

        ws = wb['BladeResult']

        # 从BladeResult工作表中读取加工信息
        # 根据常见Excel布局，重新调整单元格位置
        blade_info = {
            'blade_id': ws["B2"].value if ws["B2"].value else ws["A2"].value,  # 叶片ID
            'mill_circle_count': ws["B20"].value,  # 铣磨圈数 (通常在B3)
            'mill_depth': ws["B19"].value,  # 铣磨深度 (通常在B4)
            'start_time': ws["B4"].value,  # 加工开始时间
            'end_time': ws["B5"].value,   # 加工结束时间
            'total_duration': ws["B6"].value  # 总时长
        }

        # 关闭工作簿
        wb.close()

        return blade_info

    except Exception as e:
        raise Exception(f"读取BladeResult工作表时发生错误: {str(e)}")